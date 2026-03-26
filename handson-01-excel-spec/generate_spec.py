#!/usr/bin/env python3
"""
融資管理システム仕様書 Excel生成スクリプト
エンジニアマネージャー向けハンズオン教材用
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# スタイル定義
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_HEADER_FONT = Font(name="Yu Gothic", bold=True, size=11)
CELL_FONT = Font(name="Yu Gothic", size=10)
TITLE_FONT = Font(name="Yu Gothic", bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(name="Yu Gothic", bold=True, size=12, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_data_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGNMENT


def auto_width(ws, min_width=8, max_width=40):
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = max(length, min(max_width, len(str(cell.value)) * 1.3))
        ws.column_dimensions[col_letter].width = length


# ===========================================================================
# シート1: システム概要
# ===========================================================================
def create_system_overview(wb: Workbook):
    ws = wb.active
    ws.title = "システム概要"

    # タイトル
    ws.merge_cells("A1:F1")
    ws["A1"] = "融資管理システム (Loan Management System) 仕様書"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 基本情報テーブル
    info_header = ["項目", "内容"]
    info_data = [
        ["システム名", "融資管理システム (LoanFlow)"],
        ["バージョン", "1.0.0"],
        ["作成日", "2026-03-01"],
        ["最終更新日", "2026-03-25"],
        ["作成者", "システム開発部"],
        ["目的", "融資申請から承認・実行・返済管理までの一連のプロセスをデジタル化し、業務効率化とリスク管理強化を実現する"],
        ["対象ユーザー", "銀行の融資担当行員、支店長、審査部、システム管理者"],
        ["想定利用者数", "約500名（全国30支店）"],
        ["フロントエンド", "Next.js 14 (App Router) + TypeScript 5.x + Tailwind CSS 3.x + shadcn/ui"],
        ["バックエンド", "Next.js API Routes (Route Handlers)"],
        ["ORM", "Prisma 5.x"],
        ["データベース", "PostgreSQL 16"],
        ["認証", "NextAuth.js v5 (社内Active Directory連携)"],
        ["デプロイ", "Vercel (フロントエンド) + AWS RDS (データベース)"],
        ["監視", "Datadog APM + Sentry"],
    ]

    row = 3
    for i, h in enumerate(info_header):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT
    for r, d in enumerate(info_data, start=row + 1):
        for c, v in enumerate(d):
            cell = ws.cell(row=r, column=c + 1, value=v)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    # 機能一覧
    func_start = row + len(info_data) + 2
    ws.cell(row=func_start, column=1, value="機能一覧").font = SECTION_FONT
    func_start += 1

    func_headers = ["No", "機能名", "カテゴリ", "概要", "優先度", "ステータス"]
    for i, h in enumerate(func_headers):
        ws.cell(row=func_start, column=i + 1, value=h)
    apply_header_style(ws, func_start, len(func_headers))

    functions = [
        [1, "ユーザー認証・認可", "認証", "Active Directory連携によるSSO認証。ロールベースアクセス制御(RBAC)", "高", "必須"],
        [2, "融資申請登録", "融資管理", "ウィザード形式での融資申請データ入力。下書き保存対応", "高", "必須"],
        [3, "融資申請一覧・検索", "融資管理", "申請ステータス・金額・日付等による高度な検索・フィルタリング", "高", "必須"],
        [4, "与信審査", "審査", "申請者の信用情報・年収・勤続年数等から与信スコアを自動算出", "高", "必須"],
        [5, "承認ワークフロー", "審査", "金額に応じた多段階承認フロー（一次→二次→最終承認）", "高", "必須"],
        [6, "担保管理", "担保", "不動産・有価証券等の担保情報登録・評価・管理", "高", "必須"],
        [7, "返済スケジュール管理", "返済", "元利均等返済スケジュールの自動生成・管理・実績追跡", "高", "必須"],
        [8, "添付書類管理", "書類", "融資関連書類のアップロード・バージョン管理・電子署名", "中", "必須"],
        [9, "ダッシュボード", "分析", "融資実績KPI・承認待ち件数・滞納率等のリアルタイム表示", "中", "必須"],
        [10, "帳票出力", "帳票", "融資契約書・返済予定表・審査結果報告書のPDF出力", "中", "必須"],
        [11, "滞納管理", "返済", "返済遅延の検知・督促管理・滞納ステータス管理", "中", "Phase2"],
        [12, "外部信用情報連携", "外部連携", "CIC・JICC等の信用情報機関とのAPI連携", "中", "Phase2"],
        [13, "メール通知", "通知", "承認依頼・承認完了・返済期日リマインダーの自動メール", "低", "Phase2"],
        [14, "監査ログ", "管理", "全操作の監査証跡記録。改ざん防止ハッシュ付き", "高", "必須"],
        [15, "マスタ管理", "管理", "融資商品・金利テーブル・支店情報等のマスタデータ管理", "中", "必須"],
    ]
    for r, row_data in enumerate(functions, start=func_start + 1):
        for c, v in enumerate(row_data):
            ws.cell(row=r, column=c + 1, value=v)
        apply_data_style(ws, r, len(func_headers))

    ws.freeze_panes = "A2"


# ===========================================================================
# シート2: テーブル定義
# ===========================================================================
def create_table_definitions(wb: Workbook):
    ws = wb.create_sheet("テーブル定義")

    table_headers = ["テーブル名", "カラム名", "データ型", "NULL可否", "デフォルト値",
                     "主キー", "外部キー", "ユニーク", "インデックス", "説明"]

    tables = {
        "customers（顧客マスタ）": [
            ["customers", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "顧客ID"],
            ["customers", "customer_code", "VARCHAR(20)", "NOT NULL", "", "", "", "UK", "IDX", "顧客コード（自動採番 C-YYYYMMDD-XXXX）"],
            ["customers", "last_name", "VARCHAR(50)", "NOT NULL", "", "", "", "", "", "姓"],
            ["customers", "first_name", "VARCHAR(50)", "NOT NULL", "", "", "", "", "", "名"],
            ["customers", "last_name_kana", "VARCHAR(100)", "NOT NULL", "", "", "", "", "", "姓（カナ）"],
            ["customers", "first_name_kana", "VARCHAR(100)", "NOT NULL", "", "", "", "", "", "名（カナ）"],
            ["customers", "date_of_birth", "DATE", "NOT NULL", "", "", "", "", "", "生年月日"],
            ["customers", "gender", "VARCHAR(10)", "NOT NULL", "", "", "", "", "", "性別（male/female/other）"],
            ["customers", "phone_number", "VARCHAR(20)", "NOT NULL", "", "", "", "", "", "電話番号"],
            ["customers", "email", "VARCHAR(255)", "NULL", "", "", "", "", "", "メールアドレス"],
            ["customers", "postal_code", "VARCHAR(10)", "NOT NULL", "", "", "", "", "", "郵便番号"],
            ["customers", "address", "TEXT", "NOT NULL", "", "", "", "", "", "住所"],
            ["customers", "employer_name", "VARCHAR(200)", "NULL", "", "", "", "", "", "勤務先名"],
            ["customers", "annual_income", "DECIMAL(15,2)", "NULL", "", "", "", "", "", "年収（万円）"],
            ["customers", "employment_years", "INTEGER", "NULL", "", "", "", "", "", "勤続年数"],
            ["customers", "employment_type", "VARCHAR(30)", "NULL", "", "", "", "", "", "雇用形態（正社員/契約社員/自営業/パート）"],
            ["customers", "branch_id", "UUID", "NOT NULL", "", "", "FK→branches.id", "", "IDX", "所属支店ID"],
            ["customers", "created_by", "UUID", "NOT NULL", "", "", "FK→users.id", "", "", "登録者"],
            ["customers", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
            ["customers", "updated_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "更新日時"],
            ["customers", "deleted_at", "TIMESTAMPTZ", "NULL", "", "", "", "", "", "論理削除日時"],
        ],
        "loan_applications（融資申請）": [
            ["loan_applications", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "融資申請ID"],
            ["loan_applications", "application_number", "VARCHAR(30)", "NOT NULL", "", "", "", "UK", "IDX", "申請番号（LA-YYYYMMDD-XXXX）"],
            ["loan_applications", "customer_id", "UUID", "NOT NULL", "", "", "FK→customers.id", "", "IDX", "顧客ID"],
            ["loan_applications", "product_id", "UUID", "NOT NULL", "", "", "FK→loan_products.id", "", "IDX", "融資商品ID"],
            ["loan_applications", "requested_amount", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "申請金額（円）"],
            ["loan_applications", "approved_amount", "DECIMAL(15,2)", "NULL", "", "", "", "", "", "承認金額（円）"],
            ["loan_applications", "interest_rate", "DECIMAL(5,3)", "NULL", "", "", "", "", "", "適用金利（%）"],
            ["loan_applications", "repayment_period_months", "INTEGER", "NOT NULL", "", "", "", "", "", "返済期間（月数）"],
            ["loan_applications", "repayment_method", "VARCHAR(30)", "NOT NULL", "equal_installment", "", "", "", "", "返済方法（equal_installment/equal_principal/bullet）"],
            ["loan_applications", "purpose", "VARCHAR(50)", "NOT NULL", "", "", "", "", "", "融資目的（housing/business/education/car/other）"],
            ["loan_applications", "purpose_detail", "TEXT", "NULL", "", "", "", "", "", "融資目的詳細"],
            ["loan_applications", "status", "VARCHAR(30)", "NOT NULL", "draft", "", "", "", "IDX", "ステータス（draft/submitted/under_review/approved/rejected/executed/completed）"],
            ["loan_applications", "credit_score", "INTEGER", "NULL", "", "", "", "", "", "与信スコア（算出済みの場合）"],
            ["loan_applications", "credit_limit", "DECIMAL(15,2)", "NULL", "", "", "", "", "", "与信限度額"],
            ["loan_applications", "has_collateral", "BOOLEAN", "NOT NULL", "false", "", "", "", "", "担保有無"],
            ["loan_applications", "has_guarantor", "BOOLEAN", "NOT NULL", "false", "", "", "", "", "保証人有無"],
            ["loan_applications", "guarantor_name", "VARCHAR(100)", "NULL", "", "", "", "", "", "保証人氏名"],
            ["loan_applications", "guarantor_relationship", "VARCHAR(50)", "NULL", "", "", "", "", "", "保証人との関係"],
            ["loan_applications", "submitted_at", "TIMESTAMPTZ", "NULL", "", "", "", "", "", "申請日時"],
            ["loan_applications", "approved_at", "TIMESTAMPTZ", "NULL", "", "", "", "", "", "承認日時"],
            ["loan_applications", "executed_at", "TIMESTAMPTZ", "NULL", "", "", "", "", "", "実行日時"],
            ["loan_applications", "notes", "TEXT", "NULL", "", "", "", "", "", "備考"],
            ["loan_applications", "branch_id", "UUID", "NOT NULL", "", "", "FK→branches.id", "", "IDX", "申請支店ID"],
            ["loan_applications", "created_by", "UUID", "NOT NULL", "", "", "FK→users.id", "", "", "登録者"],
            ["loan_applications", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
            ["loan_applications", "updated_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "更新日時"],
        ],
        "loan_products（融資商品マスタ）": [
            ["loan_products", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "商品ID"],
            ["loan_products", "product_code", "VARCHAR(20)", "NOT NULL", "", "", "", "UK", "", "商品コード"],
            ["loan_products", "product_name", "VARCHAR(100)", "NOT NULL", "", "", "", "", "", "商品名（住宅ローン/ビジネスローン等）"],
            ["loan_products", "category", "VARCHAR(30)", "NOT NULL", "", "", "", "", "IDX", "カテゴリ（housing/business/education/car/free）"],
            ["loan_products", "min_amount", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "最小融資額（円）"],
            ["loan_products", "max_amount", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "最大融資額（円）"],
            ["loan_products", "base_interest_rate", "DECIMAL(5,3)", "NOT NULL", "", "", "", "", "", "基本金利（%）"],
            ["loan_products", "min_period_months", "INTEGER", "NOT NULL", "", "", "", "", "", "最短返済期間（月）"],
            ["loan_products", "max_period_months", "INTEGER", "NOT NULL", "", "", "", "", "", "最長返済期間（月）"],
            ["loan_products", "requires_collateral", "BOOLEAN", "NOT NULL", "false", "", "", "", "", "担保必須フラグ"],
            ["loan_products", "description", "TEXT", "NULL", "", "", "", "", "", "商品説明"],
            ["loan_products", "is_active", "BOOLEAN", "NOT NULL", "true", "", "", "", "", "有効フラグ"],
            ["loan_products", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
            ["loan_products", "updated_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "更新日時"],
        ],
        "collaterals（担保情報）": [
            ["collaterals", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "担保ID"],
            ["collaterals", "application_id", "UUID", "NOT NULL", "", "", "FK→loan_applications.id", "", "IDX", "融資申請ID"],
            ["collaterals", "collateral_type", "VARCHAR(30)", "NOT NULL", "", "", "", "", "", "担保種別（real_estate/securities/deposit/other）"],
            ["collaterals", "description", "TEXT", "NOT NULL", "", "", "", "", "", "担保物件の説明"],
            ["collaterals", "address", "TEXT", "NULL", "", "", "", "", "", "所在地（不動産の場合）"],
            ["collaterals", "appraised_value", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "評価額（円）"],
            ["collaterals", "appraisal_date", "DATE", "NOT NULL", "", "", "", "", "", "評価日"],
            ["collaterals", "appraiser_name", "VARCHAR(100)", "NULL", "", "", "", "", "", "評価者名"],
            ["collaterals", "coverage_ratio", "DECIMAL(5,2)", "NOT NULL", "0.70", "", "", "", "", "掛目（担保評価率）"],
            ["collaterals", "effective_value", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "有効担保額（評価額 x 掛目）"],
            ["collaterals", "registration_status", "VARCHAR(20)", "NOT NULL", "pending", "", "", "", "", "登記状況（pending/registered/released）"],
            ["collaterals", "expiry_date", "DATE", "NULL", "", "", "", "", "", "担保期限"],
            ["collaterals", "notes", "TEXT", "NULL", "", "", "", "", "", "備考"],
            ["collaterals", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
            ["collaterals", "updated_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "更新日時"],
        ],
        "credit_scores（与信スコア履歴）": [
            ["credit_scores", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "スコアID"],
            ["credit_scores", "application_id", "UUID", "NOT NULL", "", "", "FK→loan_applications.id", "", "IDX", "融資申請ID"],
            ["credit_scores", "customer_id", "UUID", "NOT NULL", "", "", "FK→customers.id", "", "IDX", "顧客ID"],
            ["credit_scores", "total_score", "INTEGER", "NOT NULL", "", "", "", "", "", "総合スコア（0-1000）"],
            ["credit_scores", "income_score", "INTEGER", "NOT NULL", "", "", "", "", "", "収入スコア（0-200）"],
            ["credit_scores", "employment_score", "INTEGER", "NOT NULL", "", "", "", "", "", "雇用安定性スコア（0-200）"],
            ["credit_scores", "credit_history_score", "INTEGER", "NOT NULL", "", "", "", "", "", "信用履歴スコア（0-200）"],
            ["credit_scores", "debt_ratio_score", "INTEGER", "NOT NULL", "", "", "", "", "", "負債比率スコア（0-200）"],
            ["credit_scores", "collateral_score", "INTEGER", "NOT NULL", "", "", "", "", "", "担保スコア（0-200）"],
            ["credit_scores", "risk_grade", "VARCHAR(5)", "NOT NULL", "", "", "", "", "IDX", "リスク等級（AAA/AA/A/BBB/BB/B/C/D）"],
            ["credit_scores", "calculated_by", "UUID", "NOT NULL", "", "", "FK→users.id", "", "", "算出者"],
            ["credit_scores", "calculated_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "算出日時"],
            ["credit_scores", "notes", "TEXT", "NULL", "", "", "", "", "", "備考・特記事項"],
        ],
        "approval_workflows（承認ワークフロー）": [
            ["approval_workflows", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "承認ID"],
            ["approval_workflows", "application_id", "UUID", "NOT NULL", "", "", "FK→loan_applications.id", "", "IDX", "融資申請ID"],
            ["approval_workflows", "step_number", "INTEGER", "NOT NULL", "", "", "", "", "", "承認ステップ番号（1:一次/2:二次/3:最終）"],
            ["approval_workflows", "step_name", "VARCHAR(50)", "NOT NULL", "", "", "", "", "", "ステップ名"],
            ["approval_workflows", "approver_id", "UUID", "NULL", "", "", "FK→users.id", "", "IDX", "承認者ID"],
            ["approval_workflows", "approver_role", "VARCHAR(30)", "NOT NULL", "", "", "", "", "", "承認者ロール"],
            ["approval_workflows", "status", "VARCHAR(20)", "NOT NULL", "pending", "", "", "", "IDX", "ステータス（pending/approved/rejected/remanded）"],
            ["approval_workflows", "comment", "TEXT", "NULL", "", "", "", "", "", "承認コメント"],
            ["approval_workflows", "conditions", "TEXT", "NULL", "", "", "", "", "", "承認条件・付帯事項"],
            ["approval_workflows", "requested_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "承認依頼日時"],
            ["approval_workflows", "decided_at", "TIMESTAMPTZ", "NULL", "", "", "", "", "", "承認判断日時"],
            ["approval_workflows", "deadline_at", "TIMESTAMPTZ", "NULL", "", "", "", "", "", "承認期限"],
            ["approval_workflows", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
        ],
        "repayment_schedules（返済スケジュール）": [
            ["repayment_schedules", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "返済スケジュールID"],
            ["repayment_schedules", "application_id", "UUID", "NOT NULL", "", "", "FK→loan_applications.id", "", "IDX", "融資申請ID"],
            ["repayment_schedules", "installment_number", "INTEGER", "NOT NULL", "", "", "", "", "", "回数"],
            ["repayment_schedules", "due_date", "DATE", "NOT NULL", "", "", "", "", "IDX", "返済期日"],
            ["repayment_schedules", "principal_amount", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "元金返済額（円）"],
            ["repayment_schedules", "interest_amount", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "利息額（円）"],
            ["repayment_schedules", "total_amount", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "返済総額（円）"],
            ["repayment_schedules", "remaining_balance", "DECIMAL(15,2)", "NOT NULL", "", "", "", "", "", "残高（円）"],
            ["repayment_schedules", "payment_status", "VARCHAR(20)", "NOT NULL", "scheduled", "", "", "", "IDX", "返済状況（scheduled/paid/overdue/partial）"],
            ["repayment_schedules", "paid_date", "DATE", "NULL", "", "", "", "", "", "実際の返済日"],
            ["repayment_schedules", "paid_amount", "DECIMAL(15,2)", "NULL", "", "", "", "", "", "実際の返済額（円）"],
            ["repayment_schedules", "overdue_days", "INTEGER", "NULL", "0", "", "", "", "", "延滞日数"],
            ["repayment_schedules", "penalty_amount", "DECIMAL(15,2)", "NULL", "0", "", "", "", "", "遅延損害金（円）"],
            ["repayment_schedules", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
        ],
        "documents（添付書類管理）": [
            ["documents", "id", "UUID", "NOT NULL", "gen_random_uuid()", "PK", "", "", "", "書類ID"],
            ["documents", "application_id", "UUID", "NOT NULL", "", "", "FK→loan_applications.id", "", "IDX", "融資申請ID"],
            ["documents", "document_type", "VARCHAR(50)", "NOT NULL", "", "", "", "", "", "書類種別（identity/income_proof/property_doc/contract/other）"],
            ["documents", "file_name", "VARCHAR(255)", "NOT NULL", "", "", "", "", "", "ファイル名"],
            ["documents", "file_path", "TEXT", "NOT NULL", "", "", "", "", "", "ファイルパス（S3 Key）"],
            ["documents", "file_size", "INTEGER", "NOT NULL", "", "", "", "", "", "ファイルサイズ（bytes）"],
            ["documents", "mime_type", "VARCHAR(100)", "NOT NULL", "", "", "", "", "", "MIMEタイプ"],
            ["documents", "version", "INTEGER", "NOT NULL", "1", "", "", "", "", "バージョン番号"],
            ["documents", "uploaded_by", "UUID", "NOT NULL", "", "", "FK→users.id", "", "", "アップロード者"],
            ["documents", "verified", "BOOLEAN", "NOT NULL", "false", "", "", "", "", "確認済みフラグ"],
            ["documents", "verified_by", "UUID", "NULL", "", "", "FK→users.id", "", "", "確認者"],
            ["documents", "created_at", "TIMESTAMPTZ", "NOT NULL", "NOW()", "", "", "", "", "登録日時"],
        ],
    }

    row = 1
    for table_label, columns in tables.items():
        ws.cell(row=row, column=1, value=table_label).font = SECTION_FONT
        row += 1
        for i, h in enumerate(table_headers):
            ws.cell(row=row, column=i + 1, value=h)
        apply_header_style(ws, row, len(table_headers))
        row += 1
        for col_data in columns:
            for c, v in enumerate(col_data):
                ws.cell(row=row, column=c + 1, value=v)
            apply_data_style(ws, row, len(table_headers))
            row += 1
        row += 1  # 空行

    auto_width(ws)
    ws.freeze_panes = "A2"


# ===========================================================================
# シート3: API仕様
# ===========================================================================
def create_api_spec(wb: Workbook):
    ws = wb.create_sheet("API仕様")

    headers = ["No", "カテゴリ", "メソッド", "エンドポイント", "概要",
               "リクエストパラメータ", "レスポンス形式", "ステータスコード", "認証", "権限"]

    apis = [
        [1, "認証", "POST", "/api/auth/login", "ログイン",
         "{ email, password }", "{ token, user, expiresAt }", "200/401/500", "不要", "全ユーザー"],
        [2, "認証", "POST", "/api/auth/logout", "ログアウト",
         "なし", "{ message }", "200/401", "Bearer", "全ユーザー"],
        [3, "融資申請", "GET", "/api/loan-applications", "融資申請一覧取得",
         "?status=&customer_id=&page=&limit=&sort=&order=", "{ data: [...], total, page, limit }", "200/401/403", "Bearer", "行員以上"],
        [4, "融資申請", "GET", "/api/loan-applications/:id", "融資申請詳細取得",
         "なし", "{ application, customer, collaterals, creditScore, workflows, documents }", "200/401/403/404", "Bearer", "行員以上"],
        [5, "融資申請", "POST", "/api/loan-applications", "融資申請新規作成",
         "{ customer_id, product_id, requested_amount, repayment_period_months, purpose, ... }", "{ application }", "201/400/401/403", "Bearer", "行員以上"],
        [6, "融資申請", "PUT", "/api/loan-applications/:id", "融資申請更新",
         "{ requested_amount?, repayment_period_months?, purpose?, ... }", "{ application }", "200/400/401/403/404", "Bearer", "行員以上"],
        [7, "融資申請", "DELETE", "/api/loan-applications/:id", "融資申請削除（下書きのみ）",
         "なし", "{ message }", "200/401/403/404/409", "Bearer", "行員以上"],
        [8, "融資申請", "POST", "/api/loan-applications/:id/submit", "融資申請提出",
         "なし", "{ application, workflow }", "200/400/401/403/409", "Bearer", "行員以上"],
        [9, "与信審査", "POST", "/api/loan-applications/:id/credit-check", "与信スコア算出",
         "なし", "{ creditScore, riskGrade, creditLimit }", "200/400/401/403/404", "Bearer", "審査部以上"],
        [10, "与信審査", "GET", "/api/loan-applications/:id/credit-scores", "与信スコア履歴取得",
         "なし", "{ scores: [...] }", "200/401/403/404", "Bearer", "行員以上"],
        [11, "承認", "POST", "/api/loan-applications/:id/approve", "融資承認",
         "{ comment?, conditions? }", "{ workflow, application }", "200/400/401/403/404/409", "Bearer", "支店長以上"],
        [12, "承認", "POST", "/api/loan-applications/:id/reject", "融資却下",
         "{ comment }", "{ workflow, application }", "200/400/401/403/404/409", "Bearer", "支店長以上"],
        [13, "承認", "POST", "/api/loan-applications/:id/remand", "差戻し",
         "{ comment, step_number? }", "{ workflow, application }", "200/400/401/403/404/409", "Bearer", "支店長以上"],
        [14, "担保", "GET", "/api/loan-applications/:id/collaterals", "担保情報一覧取得",
         "なし", "{ collaterals: [...] }", "200/401/403/404", "Bearer", "行員以上"],
        [15, "担保", "POST", "/api/loan-applications/:id/collaterals", "担保情報登録",
         "{ collateral_type, description, appraised_value, appraisal_date, coverage_ratio, ... }", "{ collateral }", "201/400/401/403/404", "Bearer", "行員以上"],
        [16, "担保", "PUT", "/api/collaterals/:id/evaluate", "担保再評価",
         "{ appraised_value, appraisal_date, appraiser_name }", "{ collateral }", "200/400/401/403/404", "Bearer", "審査部以上"],
        [17, "返済", "POST", "/api/loan-applications/:id/repayment-schedule", "返済スケジュール生成",
         "なし（申請情報から自動算出）", "{ schedule: [...], summary }", "201/400/401/403/404/409", "Bearer", "行員以上"],
        [18, "返済", "GET", "/api/loan-applications/:id/repayment-schedule", "返済スケジュール取得",
         "なし", "{ schedule: [...], summary }", "200/401/403/404", "Bearer", "行員以上"],
        [19, "返済", "PUT", "/api/repayment-schedules/:id/pay", "返済実績登録",
         "{ paid_amount, paid_date }", "{ schedule }", "200/400/401/403/404", "Bearer", "行員以上"],
        [20, "書類", "POST", "/api/loan-applications/:id/documents", "書類アップロード",
         "multipart/form-data: file, document_type", "{ document }", "201/400/401/403/404/413", "Bearer", "行員以上"],
        [21, "書類", "GET", "/api/documents/:id/download", "書類ダウンロード",
         "なし", "ファイルストリーム", "200/401/403/404", "Bearer", "行員以上"],
        [22, "ダッシュボード", "GET", "/api/dashboard/summary", "ダッシュボード集計",
         "?period=monthly&date=2026-03", "{ kpi: { totalApplications, approvedAmount, rejectionRate, avgProcessingDays, overdueRate } }", "200/401/403", "Bearer", "行員以上"],
        [23, "ダッシュボード", "GET", "/api/dashboard/pending-approvals", "承認待ち一覧",
         "?page=&limit=", "{ data: [...], total }", "200/401/403", "Bearer", "支店長以上"],
        [24, "顧客", "GET", "/api/customers", "顧客一覧取得",
         "?keyword=&page=&limit=", "{ data: [...], total, page, limit }", "200/401/403", "Bearer", "行員以上"],
        [25, "顧客", "GET", "/api/customers/:id", "顧客詳細取得",
         "なし", "{ customer, loanHistory: [...] }", "200/401/403/404", "Bearer", "行員以上"],
        [26, "顧客", "POST", "/api/customers", "顧客新規登録",
         "{ last_name, first_name, date_of_birth, phone_number, address, ... }", "{ customer }", "201/400/401/403", "Bearer", "行員以上"],
        [27, "商品", "GET", "/api/loan-products", "融資商品一覧取得",
         "?category=&is_active=true", "{ products: [...] }", "200/401", "Bearer", "全ユーザー"],
    ]

    row = 1
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, len(headers))

    for r, api in enumerate(apis, start=2):
        for c, v in enumerate(api):
            ws.cell(row=r, column=c + 1, value=v)
        apply_data_style(ws, r, len(headers))

    auto_width(ws, max_width=50)
    ws.freeze_panes = "A2"


# ===========================================================================
# シート4: 画面仕様
# ===========================================================================
def create_screen_spec(wb: Workbook):
    ws = wb.create_sheet("画面仕様")

    headers = ["No", "画面名", "URL", "機能説明", "主要表示項目", "操作・アクション",
               "バリデーションルール", "必要権限", "備考"]

    screens = [
        [1, "ログイン画面", "/login",
         "社内AD連携によるログイン。メールアドレスとパスワードで認証",
         "メールアドレス入力欄、パスワード入力欄、ログインボタン、エラーメッセージ表示領域",
         "ログインボタン押下で認証API呼出。認証成功時はダッシュボードへ遷移。失敗時はエラー表示",
         "メール: 必須・メール形式 / パスワード: 必須・8文字以上",
         "なし（未認証）",
         "3回連続失敗で30分ロック"],
        [2, "ダッシュボード", "/dashboard",
         "融資業務のKPIと承認待ち案件を一覧表示。リアルタイムで業務状況を把握可能",
         "KPIカード（今月の申請件数/承認額/却下率/平均処理日数/延滞率）、承認待ち一覧テーブル、月次推移グラフ（折れ線）、ステータス別円グラフ",
         "期間切替（月次/四半期/年次）、承認待ち案件クリックで詳細遷移、KPIカードクリックで詳細リスト表示",
         "なし（表示のみ）",
         "行員以上",
         "支店長以上は全支店データ表示。行員は自支店のみ"],
        [3, "融資申請一覧", "/loan-applications",
         "全融資申請をテーブル形式で一覧表示。多条件検索・ソート・ページネーション対応",
         "検索フォーム（ステータス/顧客名/申請番号/金額範囲/期間）、申請テーブル（申請番号/顧客名/商品名/金額/ステータス/申請日/担当者）、ページネーション",
         "新規申請ボタン、行クリックで詳細遷移、CSVエクスポート、ステータスフィルタ（タブ切替）、列ソート、ページサイズ変更（10/25/50件）",
         "金額範囲: 数値のみ・min<=max / 期間: 開始日<=終了日",
         "行員以上",
         "行員は自分の担当案件のみ。支店長は自支店全件"],
        [4, "融資申請詳細", "/loan-applications/:id",
         "融資申請の全情報をタブ切替で表示。審査・承認操作もこの画面から実行",
         "タブ構成: [基本情報] 申請番号/顧客情報/商品/金額/返済期間/ステータスバッジ / [担保] 担保一覧テーブル/評価額サマリ / [与信] スコア詳細レーダーチャート/リスク等級 / [承認履歴] ワークフロータイムライン / [書類] 添付ファイル一覧",
         "編集ボタン（下書き時のみ）、申請提出、与信審査実行、承認/差戻し/却下、担保追加、書類アップロード、返済スケジュール表示、印刷",
         "ステータスに応じた操作制限（例: 承認済みは編集不可）",
         "行員以上",
         "承認操作は権限に応じたボタン表示制御"],
        [5, "融資申請新規作成", "/loan-applications/new",
         "ウィザード形式（4ステップ）で融資申請データを入力。各ステップでリアルタイムバリデーション",
         "Step1: 顧客選択（検索・選択UI）/ Step2: 商品選択・金額入力 / Step3: 担保・保証人情報 / Step4: 確認画面。プログレスバー、戻る/次へボタン、下書き保存ボタン",
         "顧客検索・選択、商品選択、金額入力、担保追加（複数可）、保証人情報入力、書類アップロード、下書き保存、申請提出",
         "Step2: 金額は商品の範囲内/返済期間は商品の範囲内 / Step3: 担保必須商品の場合は1件以上必須 / Step4: 全必須項目入力済み確認",
         "行員以上",
         "下書き保存は何度でも可。離脱時は確認ダイアログ表示"],
        [6, "承認画面", "/loan-applications/:id/approve",
         "申請内容の確認と承認判断を行う画面。承認・差戻し・却下の3アクション",
         "申請サマリ（顧客/金額/商品/与信スコア）、承認履歴タイムライン、コメント入力欄、承認条件入力欄（任意）、承認/差戻し/却下ボタン",
         "承認ボタン（確認ダイアログ付き）、差戻しボタン（コメント必須）、却下ボタン（コメント必須・確認ダイアログ付き）",
         "差戻し・却下時はコメント必須（10文字以上）。承認時の条件は任意",
         "支店長以上（承認レベルに応じた権限チェック）",
         "金額に応じて承認レベルが異なる。権限外の案件は承認ボタン非活性"],
        [7, "顧客管理画面", "/customers",
         "顧客情報の検索・一覧表示・登録・編集。融資履歴もここから参照可能",
         "検索フォーム（氏名/顧客コード/電話番号）、顧客テーブル、登録フォーム（モーダル）、詳細パネル（融資履歴含む）",
         "新規登録、編集、検索、詳細表示、融資履歴参照、CSVエクスポート",
         "氏名: 必須 / 電話番号: 数字・ハイフンのみ / 郵便番号: XXX-XXXX形式 / メール: メール形式 / 年収: 正の数値",
         "行員以上",
         "削除は論理削除のみ。管理者権限が必要"],
        [8, "返済スケジュール画面", "/loan-applications/:id/repayment",
         "返済予定と実績を3つのビュー（カレンダー・テーブル・グラフ）で切替表示",
         "カレンダービュー（月次/返済日にマーカー表示）、テーブルビュー（回数/期日/元金/利息/合計/残高/状況）、グラフビュー（残高推移折れ線 + 元金・利息積上棒グラフ）",
         "ビュー切替（タブ）、返済実績入力（テーブルビューの各行）、繰上返済シミュレーション、PDF出力",
         "返済額: 正の数値・予定額以上 / 返済日: 期日以降",
         "行員以上",
         "繰上返済は別途承認フロー（Phase2）"],
        [9, "帳票出力画面", "/reports",
         "各種帳票のPDF出力。融資契約書・返済予定表・審査結果報告書など",
         "帳票種別選択、対象申請選択、プレビュー表示、出力ボタン",
         "帳票種別選択、対象選択、プレビュー、PDF出力、印刷",
         "対象申請は必須選択",
         "行員以上",
         "帳票テンプレートは管理者が管理"],
    ]

    row = 1
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, len(headers))

    for r, screen in enumerate(screens, start=2):
        for c, v in enumerate(screen):
            ws.cell(row=r, column=c + 1, value=v)
        apply_data_style(ws, r, len(headers))

    auto_width(ws, max_width=60)
    ws.freeze_panes = "A2"


# ===========================================================================
# シート5: ビジネスルール
# ===========================================================================
def create_business_rules(wb: Workbook):
    ws = wb.create_sheet("ビジネスルール")

    headers = ["No", "ルールID", "カテゴリ", "ルール名", "詳細説明", "計算式・条件", "適用タイミング", "例外条件"]

    rules = [
        [1, "BR-001", "与信", "与信スコア算出",
         "顧客の信用力を0-1000のスコアで算出する。5つのサブスコアの合計",
         "total_score = income_score(0-200) + employment_score(0-200) + credit_history_score(0-200) + debt_ratio_score(0-200) + collateral_score(0-200)",
         "与信審査実行時", "外部信用情報取得不可の場合はcredit_history_score=100（中立値）で算出"],
        [2, "BR-002", "与信", "収入スコア算出",
         "年収に基づくスコア。年収が高いほど高スコア",
         "年収300万未満:50 / 300-500万:100 / 500-800万:140 / 800-1200万:170 / 1200万以上:200",
         "与信スコア算出時", "自営業の場合は直近3年の平均年収を使用"],
        [3, "BR-003", "与信", "雇用安定性スコア算出",
         "雇用形態と勤続年数に基づくスコア",
         "正社員: 基礎100 + 勤続年数x10(上限100) / 契約社員: 基礎60 + 勤続年数x8(上限80) / 自営業: 基礎40 + 営業年数x10(上限100) / パート: 基礎20 + 勤続年数x5(上限40)",
         "与信スコア算出時", "公務員は正社員+20のボーナス"],
        [4, "BR-004", "融資", "融資限度額算出",
         "年収・勤続年数・信用スコアから融資可能な上限額を算出",
         "base_limit = 年収 x 倍率(スコア依存) / スコア800以上: 年収x8 / 600-799: 年収x6 / 400-599: 年収x4 / 400未満: 年収x2 / 担保ありの場合: base_limit + 有効担保額",
         "与信審査完了時・申請金額バリデーション時", "既存借入がある場合は既存残高を差し引く"],
        [5, "BR-005", "金利", "適用金利決定",
         "商品基本金利に信用スコアと担保有無による調整を加えて決定",
         "applied_rate = base_rate + score_adjustment + collateral_adjustment / score_adjustment: 800以上:-0.3% / 600-799:-0.1% / 400-599:+0.0% / 400未満:+0.5% / collateral_adjustment: 担保あり:-0.2% / 担保なし:+0.0%",
         "承認時", "キャンペーン金利が設定されている場合はそちらを優先"],
        [6, "BR-006", "返済", "元利均等返済額計算",
         "毎月の返済額（元金+利息）が均等になる返済方式",
         "monthly_payment = P * r * (1+r)^n / ((1+r)^n - 1) / P=融資額, r=月利(年利/12), n=返済回数",
         "返済スケジュール生成時", "ボーナス返済併用の場合は別途計算（Phase2）"],
        [7, "BR-007", "承認", "承認フロー分岐（金額別）",
         "融資金額に応じて必要な承認レベルが変わる",
         "500万円以下: 支店長承認のみ（1段階）/ 500万-3000万: 支店長→審査部（2段階）/ 3000万超: 支店長→審査部→本部長（3段階）",
         "申請提出時にワークフロー自動生成", "緊急案件フラグがある場合はエスカレーション可能"],
        [8, "BR-008", "承認", "承認期限",
         "各承認ステップに期限を設定。超過時はエスカレーション",
         "一次承認: 申請日から3営業日 / 二次承認: 一次承認日から5営業日 / 最終承認: 二次承認日から5営業日",
         "承認依頼時", "期限超過時は上位者+管理者にメール通知"],
        [9, "BR-009", "担保", "担保評価掛目",
         "担保種別ごとに評価額に掛目を適用して有効担保額を算出",
         "不動産（土地）: 70% / 不動産（建物）: 50% / 有価証券（上場株）: 60% / 有価証券（非上場）: 30% / 預金: 90% / その他: 個別審査",
         "担保登録・評価時", "評価額が1年以上前の場合は再評価必須"],
        [10, "BR-010", "滞納", "延滞ステータス管理",
         "返済期日からの経過日数に応じて滞納ステータスを分類",
         "1-30日: 注意（warning）/ 31-60日: 延滞（delinquent）/ 61-90日: 重度延滞（serious）/ 91日以上: 事故（default）",
         "日次バッチ処理", "自然災害等の場合は管理者により猶予期間設定可能"],
        [11, "BR-011", "滞納", "遅延損害金計算",
         "返済期日を過ぎた場合に発生する遅延損害金",
         "penalty = 延滞元金 x 遅延損害金利率(14.0%) x 延滞日数 / 365",
         "返済実績登録時・日次バッチ", "災害猶予期間中は免除"],
        [12, "BR-012", "融資", "申請ステータス遷移",
         "融資申請のステータスは定められた遷移のみ許可",
         "draft→submitted→under_review→approved→executed→completed / under_review→rejected / approved→executed（取消不可）/ どのステータスからもdraftへの差戻し可（rejected除く）",
         "各操作時", "管理者は強制ステータス変更可能（監査ログ必須）"],
        [13, "BR-013", "与信", "リスク等級判定",
         "与信スコアからリスク等級を判定",
         "900-1000: AAA / 800-899: AA / 700-799: A / 600-699: BBB / 500-599: BB / 400-499: B / 300-399: C / 300未満: D",
         "与信スコア算出時", "等級Dの場合は自動却下を推奨（最終判断は人間）"],
        [14, "BR-014", "融資", "同一顧客の重複申請チェック",
         "同一顧客が処理中の申請を複数持つことを制限",
         "同一顧客でstatus=submitted|under_reviewの申請が存在する場合、新規申請のsubmitを禁止",
         "申請提出時", "異なる商品カテゴリの場合は許可（住宅ローンとカーローンの同時申請等）"],
        [15, "BR-015", "返済", "繰上返済ルール",
         "返済期間中の繰上返済に関するルール",
         "最低繰上返済額: 10万円 / 手数料: 残高の1%（ただし融資実行から3年経過後は無料）/ 繰上返済方法: 期間短縮型 or 返済額軽減型を選択可能",
         "繰上返済申請時", "商品によっては繰上返済不可の場合あり"],
        [16, "BR-016", "与信", "負債比率スコア算出",
         "既存の借入残高と年収の比率からスコアを算出",
         "debt_ratio = 既存借入総額 / 年収 / ratio < 0.2: 200 / 0.2-0.3: 160 / 0.3-0.4: 120 / 0.4-0.5: 80 / 0.5以上: 40",
         "与信スコア算出時", "住宅ローンは負債比率計算から50%除外"],
    ]

    row = 1
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, len(headers))

    for r, rule in enumerate(rules, start=2):
        for c, v in enumerate(rule):
            ws.cell(row=r, column=c + 1, value=v)
        apply_data_style(ws, r, len(headers))

    auto_width(ws, max_width=60)
    ws.freeze_panes = "A2"


# ===========================================================================
# シート6: バリデーション定義
# ===========================================================================
def create_validation_rules(wb: Workbook):
    ws = wb.create_sheet("バリデーション定義")

    headers = ["No", "画面", "フィールド名", "項目ID", "必須", "データ型",
               "最小値/最小文字数", "最大値/最大文字数", "フォーマット",
               "クロスフィールドバリデーション", "ビジネスルール検証", "エラーメッセージ"]

    validations = [
        [1, "融資申請", "顧客", "customer_id", "必須", "UUID", "", "", "",
         "", "", "顧客を選択してください"],
        [2, "融資申請", "融資商品", "product_id", "必須", "UUID", "", "", "",
         "", "", "融資商品を選択してください"],
        [3, "融資申請", "申請金額", "requested_amount", "必須", "数値", "100000", "1000000000", "整数（円単位）",
         "選択した商品のmin_amount以上max_amount以下であること", "与信限度額以下であること（与信審査後）", "申請金額は{min}円以上{max}円以下で入力してください"],
        [4, "融資申請", "返済期間", "repayment_period_months", "必須", "整数", "1", "420", "",
         "選択した商品のmin_period_months以上max_period_months以下であること", "", "返済期間は{min}ヶ月以上{max}ヶ月以下で入力してください"],
        [5, "融資申請", "融資目的", "purpose", "必須", "選択値", "", "", "housing|business|education|car|other",
         "", "", "融資目的を選択してください"],
        [6, "融資申請", "目的詳細", "purpose_detail", "条件付必須", "テキスト", "", "1000", "",
         "purpose=otherの場合は必須", "", "融資目的の詳細を入力してください"],
        [7, "融資申請", "返済方法", "repayment_method", "必須", "選択値", "", "", "equal_installment|equal_principal|bullet",
         "", "", "返済方法を選択してください"],
        [8, "融資申請", "保証人氏名", "guarantor_name", "条件付必須", "テキスト", "2", "100", "",
         "has_guarantor=trueの場合は必須", "", "保証人氏名を入力してください"],
        [9, "顧客登録", "姓", "last_name", "必須", "テキスト", "1", "50", "",
         "", "", "姓を入力してください"],
        [10, "顧客登録", "名", "first_name", "必須", "テキスト", "1", "50", "",
         "", "", "名を入力してください"],
        [11, "顧客登録", "姓（カナ）", "last_name_kana", "必須", "テキスト", "1", "100", "全角カタカナのみ",
         "", "", "姓（カナ）を全角カタカナで入力してください"],
        [12, "顧客登録", "名（カナ）", "first_name_kana", "必須", "テキスト", "1", "100", "全角カタカナのみ",
         "", "", "名（カナ）を全角カタカナで入力してください"],
        [13, "顧客登録", "生年月日", "date_of_birth", "必須", "日付", "", "", "YYYY-MM-DD",
         "20歳以上であること", "", "生年月日を入力してください（20歳以上）"],
        [14, "顧客登録", "電話番号", "phone_number", "必須", "テキスト", "10", "15", "数字とハイフンのみ（0XX-XXXX-XXXX）",
         "", "", "電話番号を正しい形式で入力してください"],
        [15, "顧客登録", "メールアドレス", "email", "任意", "テキスト", "", "255", "RFC5322準拠",
         "", "", "正しいメールアドレス形式で入力してください"],
        [16, "顧客登録", "郵便番号", "postal_code", "必須", "テキスト", "8", "8", "XXX-XXXX",
         "", "", "郵便番号をXXX-XXXX形式で入力してください"],
        [17, "顧客登録", "住所", "address", "必須", "テキスト", "5", "500", "",
         "", "", "住所を入力してください"],
        [18, "顧客登録", "年収", "annual_income", "任意", "数値", "0", "9999999", "万円単位",
         "", "", "年収は0以上の数値で入力してください"],
        [19, "顧客登録", "勤続年数", "employment_years", "任意", "整数", "0", "60", "",
         "雇用形態が入力されている場合は必須", "", "勤続年数は0以上の整数で入力してください"],
        [20, "担保登録", "担保種別", "collateral_type", "必須", "選択値", "", "", "real_estate|securities|deposit|other",
         "", "", "担保種別を選択してください"],
        [21, "担保登録", "評価額", "appraised_value", "必須", "数値", "1", "99999999999", "円単位",
         "", "", "評価額を入力してください"],
        [22, "担保登録", "評価日", "appraisal_date", "必須", "日付", "", "", "YYYY-MM-DD",
         "現在日以前であること。1年以上前は警告表示", "", "評価日を入力してください"],
        [23, "担保登録", "掛目", "coverage_ratio", "必須", "数値", "0.01", "1.00", "小数（0.01-1.00）",
         "", "担保種別に応じたデフォルト値を自動設定", "掛目は0.01以上1.00以下で入力してください"],
        [24, "承認", "コメント", "comment", "条件付必須", "テキスト", "10", "2000", "",
         "差戻し・却下の場合は必須", "", "コメントを10文字以上で入力してください"],
        [25, "返済実績", "返済額", "paid_amount", "必須", "数値", "1", "", "円単位",
         "予定返済額以上であること", "", "返済額を入力してください"],
        [26, "返済実績", "返済日", "paid_date", "必須", "日付", "", "", "YYYY-MM-DD",
         "返済期日以降であること", "", "返済日を入力してください"],
        [27, "書類UP", "ファイル", "file", "必須", "ファイル", "", "10485760", "PDF/JPG/PNG（10MB以下）",
         "", "", "ファイルを選択してください（PDF/JPG/PNG、10MB以下）"],
        [28, "書類UP", "書類種別", "document_type", "必須", "選択値", "", "", "identity|income_proof|property_doc|contract|other",
         "", "", "書類種別を選択してください"],
    ]

    row = 1
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, len(headers))

    for r, v_rule in enumerate(validations, start=2):
        for c, v in enumerate(v_rule):
            ws.cell(row=r, column=c + 1, value=v)
        apply_data_style(ws, r, len(headers))

    auto_width(ws, max_width=50)
    ws.freeze_panes = "A2"


# ===========================================================================
# シート7: 権限マトリクス
# ===========================================================================
def create_permission_matrix(wb: Workbook):
    ws = wb.create_sheet("権限マトリクス")

    headers = ["機能", "操作", "一般行員", "支店長", "審査部", "管理者"]

    permissions = [
        ["融資申請", "一覧表示（自分の担当）", "O", "O", "O", "O"],
        ["融資申請", "一覧表示（自支店全件）", "X", "O", "O", "O"],
        ["融資申請", "一覧表示（全支店）", "X", "X", "O", "O"],
        ["融資申請", "新規作成", "O", "O", "X", "O"],
        ["融資申請", "編集（下書き）", "O", "O", "X", "O"],
        ["融資申請", "削除（下書き）", "O", "O", "X", "O"],
        ["融資申請", "申請提出", "O", "O", "X", "O"],
        ["与信審査", "スコア算出実行", "X", "X", "O", "O"],
        ["与信審査", "スコア閲覧", "O", "O", "O", "O"],
        ["承認", "一次承認（500万以下）", "X", "O", "X", "O"],
        ["承認", "二次承認（500万-3000万）", "X", "X", "O", "O"],
        ["承認", "最終承認（3000万超）", "X", "X", "X", "O"],
        ["承認", "差戻し", "X", "O", "O", "O"],
        ["承認", "却下", "X", "O", "O", "O"],
        ["担保", "登録・編集", "O", "O", "O", "O"],
        ["担保", "評価（掛目設定）", "X", "X", "O", "O"],
        ["顧客", "一覧・検索", "O", "O", "O", "O"],
        ["顧客", "新規登録", "O", "O", "X", "O"],
        ["顧客", "編集", "O", "O", "X", "O"],
        ["顧客", "論理削除", "X", "X", "X", "O"],
        ["返済スケジュール", "参照", "O", "O", "O", "O"],
        ["返済スケジュール", "生成", "O", "O", "O", "O"],
        ["返済スケジュール", "返済実績入力", "O", "O", "X", "O"],
        ["書類", "アップロード", "O", "O", "O", "O"],
        ["書類", "ダウンロード", "O", "O", "O", "O"],
        ["書類", "確認済みマーク", "X", "O", "O", "O"],
        ["ダッシュボード", "自支店データ", "O", "O", "O", "O"],
        ["ダッシュボード", "全支店データ", "X", "X", "O", "O"],
        ["帳票", "出力", "O", "O", "O", "O"],
        ["マスタ管理", "融資商品管理", "X", "X", "X", "O"],
        ["マスタ管理", "金利テーブル管理", "X", "X", "X", "O"],
        ["マスタ管理", "ユーザー管理", "X", "X", "X", "O"],
        ["監査ログ", "参照", "X", "X", "X", "O"],
    ]

    row = 1
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    apply_header_style(ws, row, len(headers))

    # O/X の色分け
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(name="Yu Gothic", bold=True, color="006100", size=10)
    red_font = Font(name="Yu Gothic", color="9C0006", size=10)

    for r, perm in enumerate(permissions, start=2):
        for c, v in enumerate(perm):
            cell = ws.cell(row=r, column=c + 1, value=v)
            cell.border = THIN_BORDER
            if c >= 2:  # 権限列
                cell.alignment = CENTER_ALIGNMENT
                if v == "O":
                    cell.fill = green_fill
                    cell.font = green_font
                else:
                    cell.fill = red_fill
                    cell.font = red_font
            else:
                cell.font = CELL_FONT
                cell.alignment = WRAP_ALIGNMENT

    # 凡例
    legend_row = len(permissions) + 3
    ws.cell(row=legend_row, column=1, value="凡例").font = SECTION_FONT
    ws.cell(row=legend_row + 1, column=1, value="O").font = green_font
    ws.cell(row=legend_row + 1, column=1).fill = green_fill
    ws.cell(row=legend_row + 1, column=2, value="操作可能").font = CELL_FONT
    ws.cell(row=legend_row + 2, column=1, value="X").font = red_font
    ws.cell(row=legend_row + 2, column=1).fill = red_fill
    ws.cell(row=legend_row + 2, column=2, value="操作不可").font = CELL_FONT

    auto_width(ws)
    ws.freeze_panes = "A2"


# ===========================================================================
# メイン実行
# ===========================================================================
def main():
    wb = Workbook()

    print("シート1: システム概要を作成中...")
    create_system_overview(wb)

    print("シート2: テーブル定義を作成中...")
    create_table_definitions(wb)

    print("シート3: API仕様を作成中...")
    create_api_spec(wb)

    print("シート4: 画面仕様を作成中...")
    create_screen_spec(wb)

    print("シート5: ビジネスルールを作成中...")
    create_business_rules(wb)

    print("シート6: バリデーション定義を作成中...")
    create_validation_rules(wb)

    print("シート7: 権限マトリクスを作成中...")
    create_permission_matrix(wb)

    output_dir = os.path.join(os.path.dirname(__file__), "specs")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "融資管理システム仕様書.xlsx")
    wb.save(output_path)
    print(f"\n仕様書を生成しました: {output_path}")
    print(f"シート数: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
